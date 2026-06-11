#!/usr/bin/env python3
"""Build a formatted Excel change-log for JDS-PRJ-GEN-001 from change-log.csv.

The CSV is the canonical, git-tracked log (append a row per change). This turns it
into a styled .xlsx — frozen header, autofilter, sensible column widths, wrapped
text — for easy review.

Usage:
    python3 scripts/build-changelog-xlsx.py
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT = Path("projects/JDS-PRJ-GEN-001")
CSV_IN = PROJECT / "change-log.csv"
XLSX_OUT = PROJECT / "change-log.xlsx"

WIDTHS = [13, 18, 26, 64, 11, 9, 11]   # per column
NAVY = "1B3A5C"


def main():
    rows = list(csv.reader(CSV_IN.open(encoding="utf-8")))
    wb = Workbook()
    ws = wb.active
    ws.title = "Change Log"

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(bold=True, color="FFFFFF")
    for r, row in enumerate(rows, 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 4))
            if r == 1:
                cell.fill = header_fill
                cell.font = header_font

    for c, width in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"

    XLSX_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUT)
    print(f"Built {XLSX_OUT} — {len(rows) - 1} rows")


if __name__ == "__main__":
    main()
