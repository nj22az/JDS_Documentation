#!/usr/bin/env python3
"""office2pdf.py — Generate JDS-compliant PDFs from Excel workbooks.

Reads timesheet, expense, or mileage Excel workbooks and renders them
as proper JDS-PRO-007 compliant PDFs using weasyprint — the same design
engine as md2pdf.py.

Unlike a naive cell-dumper, this script understands JDS document structure:
it extracts metadata, reads the data table, computes totals from actual
values (not Excel formulas), and renders everything using a purpose-built
HTML template with the JDS visual language.

Design goal: A filled timesheet should be 1–2 pages. A blank template
should be 1 page. No empty pages, no orphaned scaffolding.

Usage:
    python3 scripts/office2pdf.py <input.xlsx> [output.pdf]

Dependencies:
    pip3 install openpyxl weasyprint
"""

import sys
import os
import base64
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

try:
    from weasyprint import HTML as WeasyHTML
except ImportError:
    print("Error: weasyprint is required. Install with: pip3 install weasyprint")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(SCRIPT_DIR, '..', 'jds', 'assets')
LOGO_PATH_SVG = os.path.join(ASSETS_DIR, 'logo.svg')
LOGO_PATH_PNG = os.path.join(ASSETS_DIR, 'logo.png')
LOGO_VARIANTS_DIR = os.path.join(ASSETS_DIR, 'logo-variants')


def get_logo_data_uri(category=None):
    """Encode the JDS logo as base64 data URI. Category-coloured if available."""
    # Try category-specific SVG variant
    if category:
        variant = os.path.join(LOGO_VARIANTS_DIR, f'logo-{category.lower()}.svg')
        if os.path.exists(variant):
            with open(variant, 'r', encoding='utf-8') as f:
                data = base64.b64encode(f.read().encode('utf-8')).decode('utf-8')
            return f'data:image/svg+xml;base64,{data}'

    # Allow override via environment variable
    env_logo = os.environ.get("JDS_LOGO_PATH")
    if env_logo and os.path.exists(env_logo):
        mime = 'image/svg+xml' if env_logo.endswith('.svg') else 'image/png'
        mode = 'r' if mime == 'image/svg+xml' else 'rb'
        with open(env_logo, mode) as f:
            raw = f.read()
            if isinstance(raw, str):
                raw = raw.encode('utf-8')
            data = base64.b64encode(raw).decode('utf-8')
        return f'data:{mime};base64,{data}'

    # Default SVG
    if os.path.exists(LOGO_PATH_SVG):
        with open(LOGO_PATH_SVG, 'r', encoding='utf-8') as f:
            data = base64.b64encode(f.read().encode('utf-8')).decode('utf-8')
        return f'data:image/svg+xml;base64,{data}'

    # Fallback PNG
    if os.path.exists(LOGO_PATH_PNG):
        with open(LOGO_PATH_PNG, 'rb') as f:
            data = base64.b64encode(f.read()).decode('utf-8')
        return f'data:image/png;base64,{data}'

    return None


# ---------------------------------------------------------------------------
# Excel Data Extraction
# ---------------------------------------------------------------------------

def detect_doc_type(wb):
    """Detect document type from sheet names."""
    names = [s.lower() for s in wb.sheetnames]
    if any('timesheet' in n for n in names):
        return 'timesheet'
    elif any('expense' in n for n in names):
        return 'expense'
    elif any('mileage' in n for n in names):
        return 'mileage'
    return 'unknown'


def extract_metadata(ws, max_rows=12):
    """Extract metadata key-value pairs from the header block.

    Looks for cells containing label patterns like 'Doc No:', 'Rev:', etc.
    Returns a dict of metadata fields.
    """
    meta = {}
    label_map = {
        'doc no': 'doc_no',
        'document no': 'doc_no',
        'rev': 'revision',
        'revision': 'revision',
        'date': 'date',
        'author': 'author',
        'client': 'client',
        'project': 'project',
        'period': 'period',
        'vehicle': 'vehicle',
        'status': 'status',
    }

    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, 6):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if not isinstance(cell_val, str):
                continue
            # Only match cells that end with ':' — this distinguishes
            # metadata labels ("Date:") from table headers ("Date")
            stripped = cell_val.strip()
            if not stripped.endswith(':'):
                continue
            cell_clean = stripped.rstrip(':').strip().lower()
            if cell_clean in label_map:
                key = label_map[cell_clean]
                # Value is in the next column(s)
                val = ws.cell(row=row_idx, column=col_idx + 1).value
                if val is None:
                    # Try column after that (merged cells)
                    val = ws.cell(row=row_idx, column=col_idx + 2).value
                if val is not None:
                    meta[key] = str(val).strip()
    return meta


def find_header_row(ws, max_rows=20):
    """Find the data table header row (the row with column headers like Date, Day, etc.)."""
    header_keywords = {
        'date', 'day', 'project', 'activity', 'hours', 'category',
        'description', 'amount', 'from', 'to', 'purpose', 'km', 'rate',
        'ot', 'notes', 'vat', 'total',
    }
    for row_idx in range(1, max_rows + 1):
        row_vals = []
        for col_idx in range(1, 12):
            v = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(v, str):
                row_vals.append(v.strip().lower())
        matches = sum(1 for v in row_vals if any(k in v for k in header_keywords))
        if matches >= 3:
            return row_idx
    return None


def extract_data_table(ws, header_row):
    """Extract column headers and data rows from the worksheet."""
    # Read headers
    headers = []
    max_col = ws.max_column or 10
    for col_idx in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=col_idx).value
        if v is not None:
            headers.append(str(v).strip())
        else:
            break
    num_cols = len(headers)

    # Read data rows until we hit an empty stretch or summary keywords
    rows = []
    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        row_data = []
        has_content = False
        is_summary = False
        for col_idx in range(1, num_cols + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                s = str(v).strip()
                if s.startswith('='):
                    # Formula — try to evaluate SUM
                    v = None
                elif s.lower() in ('total hours:', 'total overtime:', 'grand total:',
                                   'subtotal:', 'total vat:', 'total distance:',
                                   'total amount:', 'total:'):
                    is_summary = True
                    break
                else:
                    has_content = True
            row_data.append(v)

        if is_summary:
            break
        if has_content:
            rows.append(row_data)

    return headers, rows


def compute_totals_timesheet(rows, headers):
    """Compute total hours and overtime from timesheet data."""
    h_idx = next((i for i, h in enumerate(headers) if h.lower() == 'hours'), None)
    ot_idx = next((i for i, h in enumerate(headers) if h.lower() == 'ot'), None)

    total_hours = 0.0
    total_ot = 0.0
    for row in rows:
        if h_idx is not None and h_idx < len(row) and row[h_idx] is not None:
            try:
                total_hours += float(row[h_idx])
            except (ValueError, TypeError):
                pass
        if ot_idx is not None and ot_idx < len(row) and row[ot_idx] is not None:
            try:
                total_ot += float(row[ot_idx])
            except (ValueError, TypeError):
                pass

    return {
        'Total Hours': f'{total_hours:.1f}',
        'Total Overtime': f'{total_ot:.1f}',
        'Grand Total': f'{total_hours + total_ot:.1f}',
    }


def compute_totals_expense(rows, headers):
    """Compute totals from expense data."""
    amt_idx = next((i for i, h in enumerate(headers) if h.lower() == 'amount'), None)
    vat_idx = next((i for i, h in enumerate(headers) if h.lower() == 'vat'), None)
    total_idx = next((i for i, h in enumerate(headers) if h.lower() == 'total'), None)

    subtotal = 0.0
    total_vat = 0.0
    grand = 0.0
    for row in rows:
        if amt_idx is not None and amt_idx < len(row) and row[amt_idx] is not None:
            try:
                subtotal += float(row[amt_idx])
            except (ValueError, TypeError):
                pass
        if vat_idx is not None and vat_idx < len(row) and row[vat_idx] is not None:
            try:
                total_vat += float(row[vat_idx])
            except (ValueError, TypeError):
                pass
        if total_idx is not None and total_idx < len(row) and row[total_idx] is not None:
            try:
                grand += float(row[total_idx])
            except (ValueError, TypeError):
                pass

    return {
        'Subtotal': f'{subtotal:,.2f} SEK',
        'Total VAT': f'{total_vat:,.2f} SEK',
        'Grand Total': f'{grand:,.2f} SEK',
    }


def compute_totals_mileage(rows, headers):
    """Compute totals from mileage data."""
    km_idx = next((i for i, h in enumerate(headers) if h.lower() == 'km'), None)
    amt_idx = next((i for i, h in enumerate(headers) if h.lower() == 'amount'), None)

    total_km = 0.0
    total_amt = 0.0
    for row in rows:
        if km_idx is not None and km_idx < len(row) and row[km_idx] is not None:
            try:
                total_km += float(row[km_idx])
            except (ValueError, TypeError):
                pass
        if amt_idx is not None and amt_idx < len(row) and row[amt_idx] is not None:
            try:
                total_amt += float(row[amt_idx])
            except (ValueError, TypeError):
                pass

    return {
        'Total Distance': f'{total_km:,.1f} km',
        'Total Amount': f'{total_amt:,.2f} SEK',
    }


# ---------------------------------------------------------------------------
# Format helpers
# ---------------------------------------------------------------------------

def fmt_cell(value, col_header=''):
    """Format a cell value for HTML display."""
    if value is None:
        return ''
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f'{value:.2f}'
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    s = str(value)
    if s.startswith('='):
        return ''
    return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def cell_align(col_header):
    """Determine alignment class for a column based on its header."""
    h = col_header.lower()
    if h in ('hours', 'ot', 'amount', 'vat', 'vat %', 'total', 'km', 'rate',
             'distance', 'count', 'excl. vat', 'incl. vat'):
        return 'num'
    if h in ('date', 'day'):
        return 'center'
    return ''


# ---------------------------------------------------------------------------
# HTML / CSS Template
# ---------------------------------------------------------------------------

def build_css(doc_no, revision, doc_date):
    """Build the JDS-compliant CSS — matches md2pdf.py design language."""
    return f"""
/* ═══════════════════════════════════════════════════════════════════════════
   JDS Office Document PDF — PRO-007 Compliant
   Same design language as md2pdf.py
   ═══════════════════════════════════════════════════════════════════════════ */

@page {{
    size: A4 portrait;
    margin: 20mm 20mm 22mm 20mm;

    @top-left {{
        content: "{doc_no}";
        font-size: 7.5pt;
        font-weight: 600;
        color: #1B3A5C;
        font-family: 'Noto Sans', 'Inter', 'Calibri', sans-serif;
        letter-spacing: 0.5pt;
        border-bottom: 0.5pt solid #e8ecf0;
        padding-bottom: 6pt;
    }}
    @top-center {{
        content: "";
        border-bottom: 0.5pt solid #e8ecf0;
        padding-bottom: 6pt;
    }}
    @top-right {{
        content: "UNCONTROLLED COPY";
        font-size: 6.5pt;
        color: #bbb;
        font-family: 'Noto Sans', 'Inter', 'Calibri', sans-serif;
        letter-spacing: 0.3pt;
        text-transform: uppercase;
        border-bottom: 0.5pt solid #e8ecf0;
        padding-bottom: 6pt;
    }}

    @bottom-left {{
        content: "Rev {revision}";
        font-size: 7pt;
        color: #999;
        font-family: 'Noto Sans', 'Inter', 'Calibri', sans-serif;
        border-top: 0.5pt solid #e8ecf0;
        padding-top: 6pt;
    }}
    @bottom-center {{
        content: "Page " counter(page) " of " counter(pages);
        font-size: 7.5pt;
        color: #999;
        font-family: 'Noto Sans', 'Inter', 'Calibri', sans-serif;
        border-top: 0.5pt solid #e8ecf0;
        padding-top: 6pt;
    }}
    @bottom-right {{
        content: "{doc_date}";
        font-size: 7pt;
        color: #999;
        font-family: 'Noto Sans', 'Inter', 'Calibri', sans-serif;
        border-top: 0.5pt solid #e8ecf0;
        padding-top: 6pt;
    }}
}}

body {{
    font-family: 'Noto Sans', 'Inter', 'Calibri', sans-serif;
    font-size: 9pt;
    line-height: 1.4;
    color: #1d1d1f;
    margin: 0;
    padding: 0;
}}

/* ═══ LOGO HEADER BAND ═══ */
.header-band {{
    margin-bottom: 4pt;
    line-height: 1;
}}
.header-band img {{
    width: 48pt;
    height: 48pt;
    vertical-align: middle;
    margin-right: 8pt;
}}
.header-band .brand {{
    font-size: 7.5pt;
    color: #86868b;
    letter-spacing: 1.5pt;
    text-transform: uppercase;
    font-weight: 600;
    vertical-align: middle;
}}

/* ═══ DOCUMENT TITLE ═══ */
h1 {{
    font-size: 20pt;
    font-weight: 700;
    color: #1B3A5C;
    border-bottom: 2.5pt solid #1B3A5C;
    padding-bottom: 6pt;
    margin: 2pt 0 8pt 0;
    letter-spacing: -0.3pt;
    line-height: 1.2;
}}

/* ═══ METADATA CARD ═══ */
.meta-card {{
    background: #fafbfc;
    border: 1pt solid #e8ecf0;
    border-radius: 8pt;
    padding: 0;
    margin: 0 0 14pt 0;
    font-size: 8.5pt;
    width: auto;
    max-width: 55%;
    overflow: hidden;
}}
.meta-card table {{
    width: 100%;
    border-collapse: collapse;
    margin: 0;
    border: none;
    border-radius: 0;
}}
.meta-card td {{
    padding: 4pt 12pt;
    border: none;
    border-bottom: 0.5pt solid #eef1f4;
    vertical-align: top;
    font-size: 8.5pt;
}}
.meta-card tr:last-child td {{
    border-bottom: none;
}}
.meta-card .label {{
    color: #86868b;
    font-weight: 600;
    text-transform: uppercase;
    font-size: 7pt;
    letter-spacing: 0.5pt;
    width: 30%;
    white-space: nowrap;
}}
.meta-card .value {{
    color: #1d1d1f;
    font-weight: 500;
}}

/* ═══ DATA TABLE ═══ */
.data-table {{
    width: 100%;
    border-collapse: separate;
    border-spacing: 0;
    margin: 0 0 12pt 0;
    font-size: 8pt;
    border: 1pt solid #e2e6ea;
    border-radius: 8pt;
    overflow: hidden;
}}
.data-table th {{
    background-color: #f5f7f9;
    color: #1B3A5C;
    font-weight: 600;
    font-size: 7.5pt;
    text-align: left;
    text-transform: uppercase;
    letter-spacing: 0.3pt;
    padding: 6pt 6pt;
    border-bottom: 1pt solid #e2e6ea;
    border-left: none;
    border-right: none;
    border-top: none;
    white-space: nowrap;
}}
.data-table td {{
    padding: 3.5pt 6pt;
    border-bottom: 0.5pt solid #eef1f4;
    border-left: none;
    border-right: none;
    vertical-align: middle;
    line-height: 1.3;
}}
.data-table tr:nth-child(even) {{
    background-color: #fafbfc;
}}
.data-table tr:last-child td {{
    border-bottom: none;
}}
.data-table tr.weekend {{
    color: #999;
}}
.data-table tr.weekend td {{
    font-style: italic;
}}
.data-table th:first-child {{ border-top-left-radius: 7pt; }}
.data-table th:last-child {{ border-top-right-radius: 7pt; }}
.data-table tr:last-child td:first-child {{ border-bottom-left-radius: 7pt; }}
.data-table tr:last-child td:last-child {{ border-bottom-right-radius: 7pt; }}

/* Numeric alignment */
.data-table .num {{
    text-align: right;
    font-variant-numeric: tabular-nums;
}}
.data-table .center {{
    text-align: center;
}}

/* ═══ TOTALS BAR ═══ */
.totals {{
    margin: 0 0 14pt 0;
    display: flex;
    gap: 0;
}}
.totals-table {{
    margin-left: auto;
    border-collapse: separate;
    border-spacing: 0;
    border: 1pt solid #1B3A5C;
    border-radius: 8pt;
    overflow: hidden;
    font-size: 8.5pt;
}}
.totals-table td {{
    padding: 5pt 14pt;
    border-bottom: 0.5pt solid #e2e6ea;
    border-left: none;
    border-right: none;
}}
.totals-table tr:last-child td {{
    border-bottom: none;
    background: #1B3A5C;
    color: white;
    font-weight: 700;
}}
.totals-table .tot-label {{
    color: #1B3A5C;
    font-weight: 600;
    text-align: right;
    padding-right: 16pt;
}}
.totals-table .tot-value {{
    text-align: right;
    font-weight: 600;
    font-variant-numeric: tabular-nums;
    min-width: 60pt;
}}
.totals-table tr:last-child .tot-label {{
    color: white;
}}

/* ═══ NOTES ═══ */
.footer-note {{
    font-size: 7.5pt;
    color: #999;
    margin-top: 8pt;
    font-style: italic;
}}
"""


def build_html(doc_type, title, meta, headers, rows, totals, logo_uri):
    """Build the complete HTML document."""
    doc_no = meta.get('doc_no', '')
    revision = meta.get('revision', 'DRAFT')
    doc_date = meta.get('date', '')
    css = build_css(doc_no, revision, doc_date)

    # Logo header band
    logo_html = ''
    if logo_uri:
        logo_html = (
            '<div class="header-band">'
            f'<img src="{logo_uri}" alt="Logo">'
            '<span class="brand">Johansson Engineering</span>'
            '</div>'
        )

    # Title
    title_html = f'<h1>{title}</h1>'

    # Metadata card
    meta_fields = [
        ('Document No.', meta.get('doc_no', 'JDS-...')),
        ('Revision', meta.get('revision', 'DRAFT')),
        ('Date', meta.get('date', '')),
        ('Author', meta.get('author', '')),
    ]
    # Add type-specific fields
    if doc_type == 'timesheet':
        if meta.get('client'):
            meta_fields.append(('Client', meta['client']))
        if meta.get('project'):
            meta_fields.append(('Project', meta['project']))
        if meta.get('period'):
            meta_fields.append(('Period', meta['period']))
    elif doc_type == 'expense':
        if meta.get('client'):
            meta_fields.append(('Client', meta['client']))
        if meta.get('project'):
            meta_fields.append(('Project', meta['project']))
        if meta.get('period'):
            meta_fields.append(('Period', meta['period']))
    elif doc_type == 'mileage':
        if meta.get('vehicle'):
            meta_fields.append(('Vehicle', meta['vehicle']))
        if meta.get('period'):
            meta_fields.append(('Period', meta['period']))

    meta_rows = ''.join(
        f'<tr><td class="label">{k}</td><td class="value">{v}</td></tr>'
        for k, v in meta_fields if v
    )
    meta_html = f'<div class="meta-card"><table>{meta_rows}</table></div>'

    # Data table
    # Filter out rows that are completely empty
    filtered_rows = []
    for row in rows:
        if any(v is not None for v in row):
            filtered_rows.append(row)

    th_html = ''.join(
        f'<th class="{cell_align(h)}">{h}</th>' for h in headers
    )

    # Detect weekend rows for timesheets
    day_idx = next((i for i, h in enumerate(headers) if h.lower() == 'day'), None)

    tr_html = ''
    for row in filtered_rows:
        is_weekend = False
        if day_idx is not None and day_idx < len(row):
            day_val = str(row[day_idx] or '').strip()
            if day_val in ('Sat', 'Sun'):
                is_weekend = True

        # Skip rows with no meaningful data (only date and day filled)
        meaningful = False
        for i, v in enumerate(row):
            if v is not None and i != 0 and (day_idx is None or i != day_idx):
                meaningful = True
                break

        if not meaningful and is_weekend:
            # Still show weekends but mark them
            pass

        row_class = ' class="weekend"' if is_weekend and not meaningful else ''
        cells = ''
        for i, v in enumerate(row):
            align_class = cell_align(headers[i]) if i < len(headers) else ''
            cls = f' class="{align_class}"' if align_class else ''
            cells += f'<td{cls}>{fmt_cell(v, headers[i] if i < len(headers) else "")}</td>'
        tr_html += f'<tr{row_class}>{cells}</tr>\n'

    table_html = f'<table class="data-table"><thead><tr>{th_html}</tr></thead><tbody>{tr_html}</tbody></table>'

    # Totals
    totals_html = ''
    if totals:
        items = list(totals.items())
        tot_rows = ''
        for label, value in items:
            tot_rows += f'<tr><td class="tot-label">{label}</td><td class="tot-value">{value}</td></tr>'
        totals_html = f'<div class="totals"><table class="totals-table">{tot_rows}</table></div>'

    # Footer note
    footer = '<p class="footer-note">UNCONTROLLED COPY — Git repository is the controlled copy per JDS-PRO-005 §6</p>'

    # Assemble
    body = f'{logo_html}{title_html}{meta_html}{table_html}{totals_html}{footer}'

    return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="utf-8"><style>{css}</style></head>
<body>{body}</body>
</html>"""


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def xlsx_to_pdf(xlsx_path, pdf_path=None):
    """Convert a JDS Excel workbook to a JDS-compliant PDF."""
    if pdf_path is None:
        pdf_path = os.path.splitext(xlsx_path)[0] + '.pdf'

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    doc_type = detect_doc_type(wb)

    # Use the first (main) sheet
    ws = wb[wb.sheetnames[0]]

    # Extract data
    meta = extract_metadata(ws)
    header_row = find_header_row(ws)
    if header_row is None:
        print(f"Error: Could not find data table header in {xlsx_path}")
        sys.exit(1)

    headers, rows = extract_data_table(ws, header_row)

    # Compute totals
    if doc_type == 'timesheet':
        title = 'Timesheet'
        totals = compute_totals_timesheet(rows, headers)
        category = 'TSH'
    elif doc_type == 'expense':
        title = 'Expense Report'
        totals = compute_totals_expense(rows, headers)
        category = 'EXP'
    elif doc_type == 'mileage':
        title = 'Mileage Log'
        totals = compute_totals_mileage(rows, headers)
        category = 'EXP'
    else:
        title = 'Document'
        totals = {}
        category = None

    logo_uri = get_logo_data_uri(category=category)
    html = build_html(doc_type, title, meta, headers, rows, totals, logo_uri)

    WeasyHTML(string=html).write_pdf(pdf_path)
    print(f"PDF saved: {pdf_path}")
    return pdf_path


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/office2pdf.py <input.xlsx> [output.pdf]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    xlsx_to_pdf(input_file, output_file)
